from flask import Flask, render_template, request, send_file, jsonify
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl import load_workbook
import os
from io import BytesIO
import tempfile

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Store template in memory (in production, use Redis or database)
template_storage = {}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload-template', methods=['POST'])
def upload_template():
    if 'template' not in request.files:
        return jsonify({'error': 'No template file provided'}), 400
    
    file = request.files['template']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        # Read and store template in memory
        template_data = file.read()
        template_storage['data'] = template_data
        template_storage['filename'] = secure_filename(file.filename)
        
        # Validate it has GA RAW sheet
        wb = load_workbook(BytesIO(template_data))
        if 'GA RAW' not in wb.sheetnames:
            return jsonify({'error': 'Template must contain "GA RAW" sheet'}), 400
        
        return jsonify({
            'message': 'Template uploaded successfully',
            'filename': template_storage['filename']
        })
    except Exception as e:
        return jsonify({'error': f'Error processing template: {str(e)}'}), 500

@app.route('/delete-template', methods=['POST'])
def delete_template():
    template_storage.clear()
    return jsonify({'message': 'Template deleted'})

@app.route('/merge', methods=['POST'])
def merge_files():
    if 'data' not in template_storage:
        return jsonify({'error': 'No template uploaded'}), 400
    
    if 'file1' not in request.files or 'file2' not in request.files:
        return jsonify({'error': 'Both data files required'}), 400
    
    file1 = request.files['file1']
    file2 = request.files['file2']
    
    try:
        # Load template from storage
        template_wb = load_workbook(BytesIO(template_storage['data']))
        
        # Load data files
        wb1 = load_workbook(BytesIO(file1.read()))
        wb2 = load_workbook(BytesIO(file2.read()))
        
        # Get first sheet from each data file
        ws1 = wb1[wb1.sheetnames[0]]
        ws2 = wb2[wb2.sheetnames[0]]
        
        # Get GA RAW sheet from template
        ga_raw = template_wb['GA RAW']
        
        # Clear existing data in GA RAW (keep the sheet structure)
        ga_raw.delete_rows(1, ga_raw.max_row)
        
        # Copy data from file1
        current_row = 1
        for row in ws1.iter_rows(values_only=True):
            for col_idx, value in enumerate(row, start=1):
                ga_raw.cell(row=current_row, column=col_idx, value=value)
            current_row += 1
        
        # Copy data from file2
        for row in ws2.iter_rows(values_only=True):
            for col_idx, value in enumerate(row, start=1):
                ga_raw.cell(row=current_row, column=col_idx, value=value)
            current_row += 1
        
        # Save to BytesIO
        output = BytesIO()
        template_wb.save(output)
        output.seek(0)
        
        # Calculate merged rows
        file1_rows = ws1.max_row
        file2_rows = ws2.max_row
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'merged_{int(os.urandom(4).hex(), 16)}.xlsx'
        ), 200, {'X-Merge-Info': f'{file1_rows},{file2_rows}'}
        
    except Exception as e:
        return jsonify({'error': f'Error merging files: {str(e)}'}), 500

@app.route('/health')
def health():
    return jsonify({'status': 'healthy'})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
